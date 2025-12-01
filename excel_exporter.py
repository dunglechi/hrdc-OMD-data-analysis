"""
Excel Exporter for VNPT Telecom Data Analysis
Creates professional Excel output with multiple sheets and formatting
"""

import pandas as pd
import json
from pathlib import Path
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# VNPT Colors
VNPT_BLUE = '0066B2'
HEADER_FILL = PatternFill(start_color=VNPT_BLUE, end_color=VNPT_BLUE, fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)


class VNPTExcelExporter:
    """Export VNPT data analysis to Excel with professional formatting"""
    
    def __init__(self, df_cleaned, stats, output_path='outputs/VNPT_Data_Analysis.xlsx'):
        """Initialize exporter"""
        self.df_cleaned = df_cleaned
        self.stats = stats
        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        
    def export_all(self):
        """Export all sheets to Excel"""
        logger.info("Exporting data to Excel...")
        
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            # Sheet 1: Executive Summary
            self._write_summary_sheet(writer)
            
            # Sheet 2: Cleaned Data
            self._write_data_sheet(writer, self.df_cleaned, 'Cleaned Data')
            
            # Sheet 3: Statistics
            self._write_statistics_sheet(writer)
            
            # Sheet 4: Customer Segments
            self._write_segments_sheet(writer)
            
            # Sheet 5: Staff Performance
            self._write_staff_sheet(writer)
            
        # Apply formatting
        self._apply_formatting()
        
        logger.info(f"Excel file saved: {self.output_path}")
        return str(self.output_path)
    
    def _write_summary_sheet(self, writer):
        """Write executive summary sheet"""
        logger.info("Writing Executive Summary sheet...")
        
        summary_data = {
            'Metric': [
                'Total Customers',
                'Service Adoption Rate',
                'High Churn Risk %',
                'Average TKC (VNĐ)',
                'Average Account Age (days)',
                'Customers with Service',
                'Customers without Service',
                'High Value Customers',
                'Unassigned Customers'
            ],
            'Value': [
                f"{self.stats['overview']['total_customers']:,}",
                f"{self.stats['service_analysis']['adoption_rate']*100:.1f}%",
                f"{self.stats['churn_analysis']['high_risk_percentage']*100:.1f}%",
                f"{self.stats['tkc_analysis']['descriptive_stats']['mean']:,.2f}",
                f"{self.stats['temporal_trends']['avg_account_age_days']:.0f}",
                f"{self.stats['service_analysis']['customers_with_service']:,}",
                f"{self.stats['service_analysis']['customers_without_service']:,}",
                f"{self.stats['segmentation']['high_value_customers']:,}",
                f"{self.stats['staff_performance']['unassigned_customers']:,}"
            ]
        }
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Executive Summary', index=False)
    
    def _write_data_sheet(self, writer, df, sheet_name):
        """Write data sheet"""
        logger.info(f"Writing {sheet_name} sheet...")
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    def _write_statistics_sheet(self, writer):
        """Write statistics sheet"""
        logger.info("Writing Statistics sheet...")
        
        # TKC Statistics
        tkc_stats = pd.DataFrame({
            'Metric': ['Mean', 'Median', 'Std Dev', 'Min', 'Max', 'Q25', 'Q75'],
            'Value': [
                self.stats['tkc_analysis']['descriptive_stats']['mean'],
                self.stats['tkc_analysis']['descriptive_stats']['median'],
                self.stats['tkc_analysis']['descriptive_stats']['std'],
                self.stats['tkc_analysis']['descriptive_stats']['min'],
                self.stats['tkc_analysis']['descriptive_stats']['max'],
                self.stats['tkc_analysis']['descriptive_stats']['q25'],
                self.stats['tkc_analysis']['descriptive_stats']['q75']
            ]
        })
        
        tkc_stats.to_excel(writer, sheet_name='Statistics', index=False, startrow=0)
        
        # TKC Segment Distribution
        segment_dist = pd.DataFrame(list(self.stats['tkc_analysis']['segment_distribution'].items()),
                                   columns=['Segment', 'Count'])
        segment_dist.to_excel(writer, sheet_name='Statistics', index=False, startrow=len(tkc_stats)+3)
    
    def _write_segments_sheet(self, writer):
        """Write customer segments sheet"""
        logger.info("Writing Customer Segments sheet...")
        
        # Parse segment matrix
        segments_data = []
        for key, value in self.stats['segmentation']['segment_matrix'].items():
            parts = key.split('_')
            tkc_segment = parts[0]
            service_status = 'With Service' if 'with_service' in key else 'No Service'
            segments_data.append({
                'TKC Segment': tkc_segment,
                'Service Status': service_status,
                'Customer Count': value['customer_count'],
                'Avg TKC': value['avg_tkc'],
                'Churn Risk Rate': f"{value['churn_risk_rate']*100:.1f}%"
            })
        
        df_segments = pd.DataFrame(segments_data)
        df_segments.to_excel(writer, sheet_name='Customer Segments', index=False)
    
    def _write_staff_sheet(self, writer):
        """Write staff performance sheet"""
        logger.info("Writing Staff Performance sheet...")
        
        # Top performers
        top_performers = self.stats['staff_performance']['top_performers']
        
        staff_data = []
        for staff_code, metrics in top_performers.items():
            if isinstance(metrics, dict):
                staff_data.append({
                    'Staff Code': staff_code,
                    'Customer Count': metrics.get('customer_count', 0),
                    'Avg TKC': metrics.get('avg_tkc', 0),
                    'Total TKC': metrics.get('total_tkc', 0),
                    'Service Rate': f"{metrics.get('service_rate', 0)*100:.1f}%"
                })
        
        df_staff = pd.DataFrame(staff_data)
        df_staff.to_excel(writer, sheet_name='Staff Performance', index=False)
    
    def _apply_formatting(self):
        """Apply professional formatting to Excel file"""
        logger.info("Applying formatting...")
        
        wb = load_workbook(self.output_path)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
        
        wb.save(self.output_path)
        logger.info("Formatting applied successfully")


def main():
    """Main execution"""
    
    logger.info("=" * 80)
    logger.info("VNPT EXCEL EXPORTER")
    logger.info("=" * 80)
    
    # Load data
    df_cleaned = pd.read_excel(r"c:\Users\Admin\.gemini\antigravity\playground\zonal-star\data\cleaned_data.xlsx")
    
    with open(r"c:\Users\Admin\.gemini\antigravity\playground\zonal-star\data\statistical_analysis.json", 'r', encoding='utf-8') as f:
        stats = json.load(f)
    
    # Export
    exporter = VNPTExcelExporter(df_cleaned, stats)
    output_file = exporter.export_all()
    
    logger.info("=" * 80)
    logger.info(f"✅ Excel export completed: {output_file}")
    logger.info("=" * 80)


if __name__ == "__main__":
    main()
